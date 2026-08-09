from __future__ import annotations

import logging
import sys
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from etl.aggregate import KEY_COLUMNS, aggregate_daily
from fg_weights import load_weights

log = logging.getLogger("update")

# -----------------------------------------------------------------------------
# CONFIG
# -----------------------------------------------------------------------------

SEASON = 2026
WINDOW_DAYS = 14
FORCE_FULL = False
MIN_PA = 1
EXCLUDE_TODAY = False
INCLUDE_FANGRAPHS = True
WOBA_SOURCE = "fangraphs"          # "fangraphs" or "statcast"
REFRESH_WEIGHTS = True
WEIGHTS_MAX_AGE_HOURS = 24
SAVE_RAW_PITCHES = False

DATA_DIR = HERE.parent / "data" / "mlb"
OUT_DIR = DATA_DIR

CHUNK_DAYS = 5
SEASON_START = (3, 1)
SEASON_END = (11, 15)

COUNT_COLS = [
    "pa", "ab", "h", "singles", "doubles", "triples", "hr",
    "bb", "ibb", "hbp", "so", "sf", "sh", "pitches",
]
RATE_COLS = [
    "avg", "obp_ish", "iso", "k_rate", "bb_rate", "woba",
    "woba_fg", "woba_savant", "pitches_per_pa",
]


def pull(start: date, end: date, use_cache: bool) -> pd.DataFrame:
    from pybaseball import cache, statcast

    if use_cache:
        cache.enable()
    else:
        try:
            cache.disable()
        except Exception:
            pass

    parts, failures = [], []
    cursor = start

    while cursor <= end:
        stop = min(cursor + timedelta(days=CHUNK_DAYS - 1), end)
        got = None

        for attempt in range(1, 4):
            try:
                got = statcast(cursor.isoformat(), stop.isoformat(), verbose=False)
                break
            except Exception as exc:
                log.warning("%s..%s attempt %d failed: %s", cursor, stop, attempt, exc)

        if got is None:
            failures.append((cursor, stop))
        elif not got.empty:
            parts.append(got)
            log.info("%s..%s  %d pitches", cursor, stop, len(got))
        else:
            log.info("%s..%s  no games", cursor, stop)

        cursor = stop + timedelta(days=1)

    if failures:
        raise RuntimeError(f"Statcast failed for {len(failures)} chunk(s): {failures}")

    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()


def merge_store(path: Path, new: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    new = new.copy()
    new["game_date"] = pd.to_datetime(new["game_date"])

    if not path.exists():
        return new.sort_values(KEY_COLUMNS).reset_index(drop=True)

    old = pd.read_parquet(path)
    old["game_date"] = pd.to_datetime(old["game_date"])
    keep = ~old["game_date"].between(pd.Timestamp(start), pd.Timestamp(end))
    combined = pd.concat([old[keep], new], ignore_index=True)

    return (
        combined
        .drop_duplicates(KEY_COLUMNS, keep="last")
        .sort_values(KEY_COLUMNS)
        .reset_index(drop=True)
    )


def div(num, den):
    den = pd.to_numeric(den, errors="coerce")
    return pd.to_numeric(num, errors="coerce") / den.where(den > 0)


def add_rates(
    frame: pd.DataFrame,
    season: int,
    woba_source: str = WOBA_SOURCE,
    weights_by_season: dict | None = None,
) -> pd.DataFrame:
    out = frame.copy()

    out["avg"] = div(out["h"], out["ab"])
    out["obp_ish"] = div(out["h"] + out["bb"] + out["hbp"], out["pa"])
    out["iso"] = div(out["doubles"] + 2 * out["triples"] + 3 * out["hr"], out["ab"])
    out["k_rate"] = div(out["so"], out["pa"])
    out["bb_rate"] = div(out["bb"], out["pa"])
    out["pitches_per_pa"] = div(out["pitches"], out["pa"])
    out["ip_est"] = out["outs_batter_events"] / 3
    out["woba_savant"] = div(out["woba_num"], out["woba_den"])

    weights = (weights_by_season or {}).get(season)
    if weights is None:
        out["woba_fg"] = np.nan
        out["woba"] = out["woba_savant"]
        log.warning("No FanGraphs wOBA weights for %d; using Savant wOBA", season)
        return out

    ubb = out["bb"] - out["ibb"]
    numerator = sum([
        weights["w1B"] * out["singles"],
        weights["w2B"] * out["doubles"],
        weights["w3B"] * out["triples"],
        weights["wHR"] * out["hr"],
        weights["wBB"] * ubb,
        weights["wHBP"] * out["hbp"],
    ])
    denominator = out["ab"] + ubb + out["hbp"] + out["sf"]

    out["woba_fg"] = div(numerator, denominator)
    out["woba"] = out["woba_fg"] if woba_source == "fangraphs" else out["woba_savant"]
    return out


def player_names() -> pd.DataFrame:
    from pybaseball import chadwick_register

    reg = chadwick_register()
    reg = reg[reg["key_mlbam"].notna()].copy()
    reg["player_id"] = pd.to_numeric(reg["key_mlbam"], errors="coerce")
    reg = reg[reg["player_id"].notna()]
    reg["player_id"] = reg["player_id"].astype("int64")
    reg["name"] = (
        reg["name_first"].fillna("").str.strip() + " " +
        reg["name_last"].fillna("").str.strip()
    ).str.strip()
    reg["key_fangraphs"] = reg["key_fangraphs"].astype("string")

    return reg[["player_id", "name", "key_fangraphs"]].drop_duplicates("player_id", keep="last")


def totals(
    store: pd.DataFrame,
    role: str,
    names: pd.DataFrame,
    hand: str | None,
    weights: dict,
) -> pd.DataFrame:
    subset = store[store["role"] == role]
    if hand:
        subset = subset[subset["opp_hand"] == hand]

    grouped = subset.groupby("player_id", as_index=False).sum(numeric_only=True)
    grouped = grouped.drop(columns="season", errors="ignore")
    out = add_rates(grouped, SEASON, WOBA_SOURCE, weights).merge(names, on="player_id", how="left")
    out = out[out["pa"] >= MIN_PA].copy()
    out["opp_hand"] = hand or "ALL"
    out["key"] = out["player_id"].astype(str) + "|" + out["opp_hand"]

    cols = ["name", "opp_hand", "player_id", "key"] + COUNT_COLS + RATE_COLS
    if role == "pitcher":
        cols.append("ip_est")
    return out[[c for c in cols if c in out]].sort_values("pa", ascending=False).reset_index(drop=True)


def fangraphs_check(ours: pd.DataFrame, role: str, names: pd.DataFrame) -> pd.DataFrame:
    from pybaseball import batting_stats, pitching_stats

    fg = (batting_stats if role == "batter" else pitching_stats)(SEASON, qual=0)
    id_col = next((c for c in ("IDfg", "playerid") if c in fg), None)
    if not id_col:
        return pd.DataFrame()

    mapping = {
        "PA": "pa", "TBF": "pa", "AB": "ab", "H": "h", "2B": "doubles",
        "3B": "triples", "HR": "hr", "BB": "bb", "IBB": "ibb", "SO": "so",
        "HBP": "hbp", "AVG": "avg", "ISO": "iso", "wOBA": "woba",
        "K%": "k_rate", "BB%": "bb_rate",
    }

    theirs = pd.DataFrame({"key_fangraphs": fg[id_col].astype(str).str.strip()})
    if "Name" in fg:
        theirs["fg_name"] = fg["Name"]

    for source, target in mapping.items():
        if source in fg:
            theirs[f"fg_{target}"] = pd.to_numeric(fg[source], errors="coerce")

    for col in ("fg_k_rate", "fg_bb_rate"):
        if col in theirs and theirs[col].dropna().median() > 1:
            theirs[col] /= 100

    merged = (
        ours
        .merge(names[["player_id", "key_fangraphs"]], on="player_id", how="left")
        .merge(theirs, on="key_fangraphs", how="left")
    )

    for stat in ["pa", "ab", "h", "doubles", "triples", "hr", "bb", "so", "hbp",
                 "avg", "iso", "woba", "k_rate", "bb_rate"]:
        if stat in merged and f"fg_{stat}" in merged:
            merged[f"d_{stat}"] = merged[stat] - merged[f"fg_{stat}"]

    front = [c for c in ("name", "fg_name", "player_id", "key_fangraphs") if c in merged]
    return merged[front + [c for c in merged if c not in front]]


def write_workbook(sheets: dict[str, pd.DataFrame], path: Path) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    formats = {
        **dict.fromkeys(("avg", "obp_ish", "iso", "woba"), "0.000"),
        **dict.fromkeys(("k_rate", "bb_rate"), "0.0%"),
        **dict.fromkeys(("pitches_per_pa", "ip_est"), "0.00"),
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            if frame is None or frame.empty:
                continue

            frame.to_excel(writer, sheet_name=name[:31], index=False)
            sheet = writer.sheets[name[:31]]
            sheet.freeze_panes = "C2"
            sheet.auto_filter.ref = sheet.dimensions

            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for i, column in enumerate(frame.columns, 1):
                letter = get_column_letter(i)
                sheet.column_dimensions[letter].width = (
                    24 if column in {"name", "fg_name"} else max(len(column) + 2, 10)
                )
                base = column.removeprefix("d_").removeprefix("fg_")
                if base not in formats:
                    continue
                for cell in sheet.iter_cols(min_col=i, max_col=i, min_row=2):
                    for item in cell:
                        item.number_format = formats[base]


def resolve_window(store_path: Path) -> tuple[date, date, bool]:
    latest = date.today() - timedelta(days=1) if EXCLUDE_TODAY else date.today()
    season_end = min(date(SEASON, *SEASON_END), latest)

    if FORCE_FULL or not store_path.exists():
        return date(SEASON, *SEASON_START), season_end, True

    return season_end - timedelta(days=WINDOW_DAYS - 1), season_end, False


def build_sheets(store: pd.DataFrame, role: str, names: pd.DataFrame, weights: dict) -> dict[str, pd.DataFrame]:
    combined = totals(store, role, names, None, weights)
    splits = pd.concat(
        [totals(store, role, names, hand, weights) for hand in ("L", "R")],
        ignore_index=True,
    ).sort_values(["name", "opp_hand"], na_position="last").reset_index(drop=True)

    sheets = {"splits": splits}
    if INCLUDE_FANGRAPHS:
        try:
            check = fangraphs_check(combined, role, names)
            if not check.empty:
                sheets["fangraphs_check"] = check
        except Exception as exc:
            log.warning("FanGraphs comparison failed for %s: %s", role, exc)
    return sheets


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)-7s %(message)s",
        stream=sys.stdout,
        force=True,
    )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    store_path = DATA_DIR / f"daily_components_{SEASON}.parquet"
    start, end, full_run = resolve_window(store_path)

    log.info("window %s .. %s (%s)", start, end, "full season" if full_run else f"last {WINDOW_DAYS} days")

    pitches = pull(start, end, use_cache=full_run)
    if pitches.empty:
        log.info("no pitches found; nothing changed")
        return

    if SAVE_RAW_PITCHES:
        pitches.to_parquet(DATA_DIR / f"pitches_{SEASON}.parquet", index=False)

    daily = aggregate_daily(pitches, game_types=("R",))
    store = merge_store(store_path, daily, start, end)
    store.to_parquet(store_path, index=False)

    weights, provenance = load_weights(
        DATA_DIR / "fg_woba_weights.json",
        SEASON,
        max_age_hours=WEIGHTS_MAX_AGE_HOURS,
        allow_network=REFRESH_WEIGHTS,
    )
    log.info("wOBA weights: %s", provenance)

    names = player_names()

    for role, label in (("batter", "hitters"), ("pitcher", "pitchers")):
        sheets = build_sheets(store, role, names, weights)
        path = OUT_DIR / f"statcast_{label}_{SEASON}.xlsx"
        write_workbook(sheets, path)
        log.info("wrote %s", path)

    log.info("update complete")


if __name__ == "__main__":
    main()
